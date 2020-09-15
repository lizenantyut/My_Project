#coding: utf8
import pandas as pd
import numpy as np
import requests
import json
from openpyxl import load_workbook
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


def test():
    """
    这里是测试方法的实例 以及是否可用
    :return:
    """
    #data_to_excel
    data_to_excel(df,"test1.xlsx","2020-09-04",
                  "D:\\lzn\\Python_Project\\知识总结\\Python\\python和Excel结合\\123")
    data_to_excel(df, "test1.xlsx", "2020-09-04")

def reset_col(filename):
    """
    根据每列的内容自动调整列宽
    :param filename:
    :return:
    """
    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        df = pd.read_excel(filename, sheet).fillna('-')
        df.loc[len(df)] = list(df.columns)  # 把标题行附加到最后一行
        for col in df.columns:
            index = list(df.columns).index(col)  # 列序号
            letter = get_column_letter(index + 1)  # 列字母
            collen = df[col].apply(lambda x: len(str(x).encode())).max()  # 获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
            ws.column_dimensions[letter].width = collen  + 2

def request_post_data(url, param):
    """
    获取数据源
    :param url:
    :param param:
    :return:
    """
    fails = 0
    while True:
        try:
            if fails >= 20:
                break
            headers = {'content-type': 'application/json'}
            ret = requests.post(url, json=param, headers=headers, timeout=10)
            if ret.status_code == 200:
                text = json.loads(ret.text)
            else:
                continue
        except:
            fails += 1
            print('网络连接出现问题, 正在尝试再次请求: ', fails)
        else:
            break
    return text

def excel_ys():
    """设置Excel样式"""
    pass


def adjust_column_width_from_col(ws, min_row, min_col, max_col):
    column_widths = []
    for i, col in \
            enumerate(
                ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)
            ):

        for cell in col:
            value = cell.value
            if value is not None:

                if isinstance(value, str) is False:
                    value = str(value)

                try:
                    column_widths[i] = max(column_widths[i], len(value))
                except IndexError:
                    column_widths.append(len(value))

    for i, width in enumerate(column_widths):
        col_name = get_column_letter(min_col + i)
        value = column_widths[i] + 2
        ws.column_dimensions[col_name].width = value



def data_to_excel(df,excle_name,sheet_name,excel_path=None):
    """

    :param df: 要插入excel的dataframe数据
    :param excle_name:  要写入的excel文件名称
    :param sheet_name:  要写入的sheet名
    :param excel_path:  excel路径 默认为None 如果不传入默认为当前路径 即程序所在路径
    :return:
    """

    #这里是对文件是否存在做判断
    if excel_path is None:
        excelWriter = pd.ExcelWriter(excle_name, engine='openpyxl')
        file_flag = os.path.exists(excle_name)
    else:
        if excel_path.endswith("\\"):
            excelWriter = pd.ExcelWriter(excel_path+excle_name, engine='openpyxl')
            file_flag = os.path.exists(excel_path+excle_name)
        else:
            excelWriter = pd.ExcelWriter(excel_path + "\\" + excle_name, engine='openpyxl')
            file_flag = os.path.exists(excel_path+"\\"+excle_name)

    #写入df到Excel文件 存在文件和不存在文件判断 避免在重新写入数据时覆盖原有sheet页
    if file_flag:
        book = load_workbook(excelWriter.path)
        excelWriter.book = book
        df.to_excel(excel_writer=excelWriter, sheet_name=sheet_name, index=None)
        excelWriter.close()
    else:
        df.to_excel(excel_writer=excelWriter, sheet_name=sheet_name, index=False)
        excelWriter.save()
        excelWriter.close()

def test2():
    wb = openpyxl.load_workbook("test1.xlsx")
    sheet = wb.get_sheet_by_name("2020-09-02")

    for col in range(1,1000):
      #col_value = data_dict[header_dict[str(col)]]
      #ws.cell(row_num, col, col_value).number_format = fmt_acct
      sheet.column_dimensions[get_column_letter(col)].auto_size = True
    wb.save("test1.xlsx")
    wb.close()
    #wb.save("test1.xlsx")



if __name__ == '__main__':
    post_url = "http://10.55.202.164:40701/epidemic/getAbnormalOnBoardWarnBusDataList"
    request_param = {"date": "2020-09-01"}
    a = request_post_data(post_url, request_param)
    js = json.dumps(a, sort_keys=True, indent=4, separators=(',', ':'))  # 格式化转json格式
    df = pd.read_json(js, orient='records')

    data_to_excel(df,"test1.xlsx","2020-09-02")


